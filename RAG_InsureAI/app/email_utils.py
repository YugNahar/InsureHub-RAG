"""
Email + PDF utilities for offline escalation.
When no agent is online and the AI can't answer, this module:
  1. Generates a PDF transcript with the unanswerable query highlighted
  2. Uses the LLM to compose a professional email body
  3. Sends the email + PDF to every agent listed in the agent-emails database
"""
import logging
import os
import re
import smtplib
from datetime import datetime, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List

logger = logging.getLogger(__name__)

import email_settings

# GMAIL_SENDER / GMAIL_APP_PASSWORD used to be frozen module-level
# constants read once at import — moved to email_settings.get_sender() /
# get_app_password() (Super Admin panel, 2026-08-14) so a credential
# rotation takes effect on the very next send, no restart needed. Every
# use below is a live call, not a cached value.
# Legacy single-recipient fallback — only used if the agent-emails database
# file itself doesn't exist yet (first-time setup). Once the file exists, it
# is the sole source of truth, even if that means it's empty.
AGENT_EMAIL     = os.getenv("AGENT_EMAIL", "lavishdevpura6@gmail.com")
VLLM_HOST       = os.getenv("VLLM_HOST", "")
VLLM_MODEL      = os.getenv("VLLM_MODEL", "")

# vLLM is the only LLM backend allowed in production for email composition
# (Groq was tried and measured far faster, but isn't an approved option on
# the production side — reverted). This deployment decodes at only ~7-8
# tokens/sec (confirmed live), so the timeout and max_tokens below are
# calibrated for THAT throughput, not for a fast backend: max_tokens is
# capped at 450 (comfortably covers the "under 250/300 words" asked for in
# each prompt below) and the timeout is set well above 450 tokens' worst-
# case generation time (~60s) so a normal-length completion has room to
# actually finish instead of racing a timeout tuned for a fast backend and
# losing almost every time. smtplib's own timeout (Gmail auth failures are
# fast, not slow) is unaffected and stays short.
_LLM_TIMEOUT_S = 75
_LLM_MAX_TOKENS = 450
_SMTP_TIMEOUT_S = 10


def _smtp_send(sender: str, password: str, recipients: List[str], msg) -> None:
    """Connects, authenticates, and sends via whichever provider is
    currently configured (email_settings.get_provider()) — Gmail wants
    SMTP_SSL on port 465; Outlook/Office 365 wants plain SMTP + STARTTLS
    on port 587. Both providers still just take a mailbox address + an
    App Password, so nothing else about the caller changes.

    Note for Outlook/Office 365 specifically: Microsoft disabled SMTP AUTH
    tenant-wide by default a few years back — if login fails with an
    authentication error even though the App Password is definitely
    correct, an admin needs to explicitly re-enable SMTP AUTH for that
    mailbox in the Microsoft 365 admin center first. Not something this
    code can detect or work around.
    """
    cfg = email_settings.PROVIDERS.get(
        email_settings.get_provider(), email_settings.PROVIDERS["gmail"]
    )
    if cfg["mode"] == "ssl":
        with smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=_SMTP_TIMEOUT_S) as server:
            server.login(sender, password)
            server.sendmail(sender, recipients, msg.as_string())
    else:  # "starttls"
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=_SMTP_TIMEOUT_S) as server:
            server.starttls()
            server.login(sender, password)
            server.sendmail(sender, recipients, msg.as_string())


def _get_composer_client():
    """vLLM only (see note above) — returns (client, model) via the shared
    remote host, or (None, None) if VLLM_HOST isn't configured. Timeout and
    retry settings are applied here so every caller gets them uniformly."""
    if not VLLM_HOST:
        return None, None
    from openai import OpenAI
    client = OpenAI(base_url=f"{VLLM_HOST}/v1", api_key="dummy", timeout=_LLM_TIMEOUT_S, max_retries=0)
    return client, VLLM_MODEL

# Excel "database" of agent recipient emails — one email per row under an
# "email" header in the first sheet. Edit this file (add/remove rows) to
# change who gets escalation emails; no restart or code change needed, since
# it's re-read from disk on every send rather than cached in memory.
AGENT_EMAILS_FILE = os.getenv(
    "AGENT_EMAILS_FILE",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "agent_emails.xlsx"),
)
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _load_agent_email_entries() -> List[dict]:
    """Read {"name", "email"} entries from AGENT_EMAILS_FILE — email in
    column A (unchanged position, so any external tooling/exports that
    only ever knew about the single-column legacy file still work), name
    in column B (new, optional — blank for legacy rows or a bulk upload
    that only had an email column).

    Re-reads the file from disk on every call (no caching) so edits take
    effect on the very next read/send without needing a restart. Returns
    [] if the file is missing, unreadable, or has no valid email rows.
    No limit on how many rows this returns — the super-admin panel and the
    send path both handle however many agents are actually in the file.
    """
    if not os.path.exists(AGENT_EMAILS_FILE):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(AGENT_EMAILS_FILE, read_only=True, data_only=True)
        ws = wb.active
        entries: List[dict] = []
        seen = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            email_cell = row[0] if len(row) > 0 else None
            name_cell = row[1] if len(row) > 1 else None
            if email_cell is None:
                continue
            email = str(email_cell).strip()
            if not email or not _EMAIL_RE.match(email):
                continue
            key = email.lower()
            if key in seen:
                continue
            seen.add(key)
            name = str(name_cell).strip() if name_cell is not None else ""
            entries.append({"name": name, "email": email})
        return entries
    except Exception:
        logger.exception("Failed to read agent-emails database at %s", AGENT_EMAILS_FILE)
        return []


def _save_agent_email_entries(entries: List[dict]) -> None:
    """Overwrite AGENT_EMAILS_FILE with the given entries — one row each,
    email in column A, name in column B. Called by the super-admin panel's
    add/remove/upload actions; the send path only ever reads afterward."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "agents"
    ws.append(["email", "name"])
    for e in entries:
        ws.append([e["email"], e.get("name", "")])
    wb.save(AGENT_EMAILS_FILE)


def _parse_email_entries_from_bytes(data: bytes, filename: str) -> List[dict]:
    """Parse an uploaded .csv or .xlsx file into [{"name", "email"}, ...].
    Looks for "email"/"name" column headers case-insensitively, in
    whatever order they appear; falls back to treating a headerless file
    as a plain single-column list of emails (column A). Raises ValueError
    for anything else (unsupported extension)."""
    ext = os.path.splitext(filename)[1].lower()
    rows: List[List[str]] = []
    if ext in (".xlsx", ".xlsm"):
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in row])
    elif ext == ".csv":
        import csv
        import io
        text = data.decode("utf-8-sig", errors="replace")
        rows = [[c.strip() for c in row] for row in csv.reader(io.StringIO(text))]
    else:
        raise ValueError("Only .csv or .xlsx files are supported")

    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return []

    header = [c.lower() for c in rows[0]]
    has_header = "email" in header
    email_col = header.index("email") if has_header else 0
    name_col = header.index("name") if "name" in header else None
    data_rows = rows[1:] if has_header else rows

    entries: List[dict] = []
    seen = set()
    for row in data_rows:
        if email_col >= len(row):
            continue
        email = row[email_col].strip()
        if not email or not _EMAIL_RE.match(email):
            continue
        key = email.lower()
        if key in seen:
            continue
        seen.add(key)
        name = row[name_col].strip() if name_col is not None and name_col < len(row) else ""
        entries.append({"name": name, "email": email})
    return entries


def _load_agent_emails() -> List[str]:
    """Back-compat wrapper for the send path — just the email addresses,
    in the same dedup'd order _load_agent_email_entries returns them."""
    return [e["email"] for e in _load_agent_email_entries()]


def _safe(text: str) -> str:
    """Strip characters outside latin-1 so fpdf2's built-in Helvetica never errors."""
    return (
        str(text)
        .replace("—", "-").replace("–", "-")   # em/en dash
        .replace("‘", "'").replace("’", "'")   # curly single quotes
        .replace("“", '"').replace("”", '"')   # curly double quotes
        .replace("…", "...")                         # ellipsis
        .replace("â", "'")              # mangled UTF-8
        .encode("latin-1", errors="replace").decode("latin-1")
    )


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf(session_id: str, history, unanswerable_query: str) -> bytes:
    from fpdf import FPDF

    class PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 13)
            self.set_fill_color(79, 70, 229)
            self.set_text_color(255, 255, 255)
            self.cell(0, 10, "  InsureHub - Unresolved Support Request", ln=True, fill=True)
            self.set_text_color(0, 0, 0)
            self.ln(2)

    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Meta
    pdf.set_font("Helvetica", size=9)
    pdf.set_text_color(100, 100, 100)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    pdf.cell(0, 6, _safe(f"Session: #{session_id}   |   Generated: {ts}"), ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # Highlighted unanswerable query box
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(254, 243, 199)
    pdf.set_draw_color(217, 119, 6)
    pdf.set_line_width(0.5)
    pdf.cell(0, 7, "  QUERY THAT REQUIRES AGENT ATTENTION", ln=True, fill=True, border=1)
    pdf.set_font("Helvetica", "I", 10)
    pdf.set_fill_color(255, 251, 235)
    pdf.multi_cell(0, 7, _safe(f'  "{unanswerable_query}"'), fill=True, border="LRB")
    pdf.set_line_width(0.2)
    pdf.set_draw_color(0, 0, 0)
    pdf.ln(8)

    # Transcript
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "FULL CONVERSATION TRANSCRIPT", ln=True)
    pdf.set_draw_color(200, 200, 200)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(4)

    uq_lower = unanswerable_query.strip().lower()
    for msg in history:
        if msg.role == "system":
            continue
        if msg.role == "user":
            is_unanswered = msg.content.strip().lower() == uq_lower
            if is_unanswered:
                pdf.set_fill_color(254, 226, 226)
                prefix = "[USER - UNANSWERED]"
            else:
                pdf.set_fill_color(239, 246, 255)
                prefix = "[USER]"
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(30, 64, 175)
            pdf.multi_cell(0, 7, _safe(f"  {prefix}  {msg.content}"), fill=True)
        elif msg.role in ("ai", "agent"):
            pdf.set_fill_color(249, 250, 251)
            pdf.set_font("Helvetica", size=9)
            pdf.set_text_color(55, 65, 81)
            label = "[LAYLA (AI)]" if msg.role == "ai" else "[AGENT]"
            pdf.multi_cell(0, 6, _safe(f"  {label}  {msg.content}"), fill=True)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    pdf.ln(4)
    pdf.set_font("Helvetica", size=8)
    pdf.set_text_color(128, 128, 128)
    pdf.cell(0, 5, "Please review the highlighted query and follow up with the user at your earliest convenience.", ln=True)

    return bytes(pdf.output())


# ── LLM email composition ─────────────────────────────────────────────────────

def compose_email_body(session_id: str, history, unanswerable_query: str) -> str:
    """Writes a professional HTML email body (Groq preferred, vLLM fallback —
    see _get_composer_client). Falls back to template if neither is configured
    or the call fails."""
    client, model = _get_composer_client()
    if client is None:
        return _template_email(session_id, unanswerable_query, history)

    history_lines = []
    for m in history:
        if m.role == "user":
            history_lines.append(f"User: {m.content}")
        elif m.role == "ai":
            history_lines.append(f"AI: {m.content[:300]}")
    conversation_text = "\n".join(history_lines[-20:])   # last 20 turns

    prompt = f"""You are writing a professional support escalation email from an AI insurance assistant.

CONTEXT:
- Platform: InsureHub AI Insurance Advisor ("Layla")
- Session: #{session_id}
- A user asked a question Layla could not answer
- No human agent was available so this email is auto-generated
- A PDF transcript is attached

THE UNANSWERABLE QUESTION (must be highlighted in the email):
"{unanswerable_query}"

CONVERSATION SUMMARY (last few turns):
{conversation_text}

Write a concise professional HTML email body to the support agent. Requirements:
- Open with a brief intro (1-2 sentences)
- Include the unanswerable question in bold/highlighted HTML
- Give a 2-3 sentence summary of the conversation context
- Mention the PDF transcript is attached
- End with a polite request to follow up with the user promptly
- Keep it under 250 words
- HTML only — no subject line, no To/From headers, just the body content
- Use simple tags: <p>, <b>, <span style="background:#fef3c7;padding:2px 6px">, <ul>, <li>"""

    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=_LLM_MAX_TOKENS,
            temperature=0.4,
        )
        body = resp.choices[0].message.content.strip()
        # Ensure it's wrapped in basic HTML if not already
        if not body.strip().startswith("<"):
            body = f"<p>{body}</p>"
        return body
    except Exception as e:
        logger.warning("LLM email composition failed: %s — using template", e)
        return _template_email(session_id, unanswerable_query, history)


def _template_email(session_id: str, unanswerable_query: str, history) -> str:
    msg_count = sum(1 for m in history if m.role == "user")
    return f"""
<p>Hello,</p>

<p>A user on <b>InsureHub</b> needed help that our AI assistant Layla could not provide.
No agent was online at the time, so this email is being sent automatically.</p>

<p><b>Unanswerable Query:</b></p>
<p style="background:#fef3c7;padding:10px 14px;border-left:4px solid #d97706;border-radius:4px">
  &ldquo;{unanswerable_query}&rdquo;
</p>

<p>The user had {msg_count} message(s) in this session (Session ID: <b>#{session_id}</b>).
The full conversation transcript is attached as a PDF for your review.</p>

<p>Please follow up with the user at your earliest convenience.</p>

<p>Best regards,<br>
<b>InsureHub AI System</b></p>
"""


# ── Email dispatch ────────────────────────────────────────────────────────────

def send_escalation_email(session_id: str, history, unanswerable_query: str) -> bool:
    """
    Generate PDF + compose email body with LLM + send via SMTP (Gmail or
    Outlook/Office 365, whichever is configured) to every agent listed in
    the agent-emails database (AGENT_EMAILS_FILE).

    No per-agent credentials are needed — the system sends FROM one shared
    sender account (configurable in the Super Admin panel) TO every address
    the database currently lists, so adding or removing a row there changes
    who receives escalations without touching any code or credentials.

    Returns True on success, False on failure or if there are no recipients
    (logs the reason either way).
    """
    sender, password = email_settings.get_sender(), email_settings.get_app_password()
    if not sender or not password:
        logger.warning(
            "Email escalation skipped — no sender/app password configured. "
            "Set these in Super Admin > Settings > Escalation Email Sender "
            "(or GMAIL_SENDER/GMAIL_APP_PASSWORD or OUTLOOK_SENDER/OUTLOOK_APP_PASSWORD in .env)."
        )
        return False

    recipients = _load_agent_emails()
    if not recipients and not os.path.exists(AGENT_EMAILS_FILE):
        # Database file doesn't exist yet — fall back to the legacy single
        # recipient so a fresh install isn't silently mute before anyone's
        # created the file.
        recipients = [AGENT_EMAIL] if AGENT_EMAIL else []
    if not recipients:
        logger.warning(
            "Email escalation skipped for session %s — agent-emails database "
            "at %s has no valid recipients.", session_id, AGENT_EMAILS_FILE,
        )
        return False

    try:
        pdf_bytes = generate_pdf(session_id, history, unanswerable_query)
        body_html = compose_email_body(session_id, history, unanswerable_query)

        msg = MIMEMultipart("mixed")
        msg["From"]    = sender
        # Real recipients go in Bcc (the SMTP envelope below, not a visible
        # header) so agents don't see each other's addresses. "To" needs
        # *some* address to be a well-formed message — use the sender's own.
        msg["To"]      = sender
        msg["Subject"] = f"[InsureHub] User Needs Help — Session #{session_id}"

        # HTML body
        alt = MIMEMultipart("alternative")
        plain = f"A user (session #{session_id}) had an unanswerable question: {unanswerable_query}\nSee attached PDF for full transcript."
        alt.attach(MIMEText(plain, "plain"))
        alt.attach(MIMEText(_email_wrapper(body_html), "html"))
        msg.attach(alt)

        # PDF attachment
        part = MIMEApplication(pdf_bytes, _subtype="pdf")
        part.add_header("Content-Disposition", "attachment",
                        filename=f"insurehub_session_{session_id}.pdf")
        msg.attach(part)

        _smtp_send(sender, password, recipients, msg)

        logger.info(
            "Escalation email sent for session %s → %d agent(s): %s",
            session_id, len(recipients), ", ".join(recipients),
        )
        return True

    except Exception:
        logger.exception("Failed to send escalation email for session %s", session_id)
        return False


# ── Quotation email (Ava, travel_bot) ───────────────────────────────────────
# Sent on-demand from the "Email me this quote" button in the travel-bot
# chat, not part of the offline-escalation flow above — reuses the same
# vLLM-composition-with-template-fallback and SMTP-send shape, but goes
# directly to the customer's own address (no Bcc-to-agents trick, no PDF).

def compose_quotation_email_body(session_id: str, recipient_name: str, trip_summary: str, quotes_text: str) -> str:
    """Writes a polished quotation summary email (Groq preferred, vLLM
    fallback — see _get_composer_client). Falls back to a plain template if
    neither is configured or the call fails — sending should never block on
    the LLM being unavailable."""
    client, model = _get_composer_client()
    if client is None:
        return _template_quotation_email(recipient_name, trip_summary, quotes_text)

    prompt = f"""You are writing a polished quotation email from a travel-insurance
assistant ("Ava", part of the InsureHub platform) to a customer who just
requested their travel insurance quote(s) by email.

CUSTOMER NAME: {recipient_name or "there"}
SESSION: #{session_id}

TRIP DETAILS:
{trip_summary}

QUOTE OPTIONS:
{quotes_text}

Write a warm, professional HTML email body. Requirements:
- Greet the customer by name (or "there" if no name given)
- Briefly restate the trip details in one short line
- Present the quote options clearly (an HTML table or a clean list is fine —
  keep the exact numbers/insurer/plan names from QUOTE OPTIONS above, do not
  invent or alter any figures)
- Close with a friendly note that they can reply with any questions
- Keep it under 300 words
- HTML only — no subject line, no To/From headers, just the body content
- Use simple tags: <p>, <b>, <table>, <tr>, <td>, <ul>, <li>"""

    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=_LLM_MAX_TOKENS,
            temperature=0.4,
        )
        body = resp.choices[0].message.content.strip()
        if not body.strip().startswith("<"):
            body = f"<p>{body}</p>"
        return body
    except Exception as e:
        logger.warning("LLM quotation-email composition failed: %s — using template", e)
        return _template_quotation_email(recipient_name, trip_summary, quotes_text)


def _template_quotation_email(recipient_name: str, trip_summary: str, quotes_text: str) -> str:
    quotes_html = "".join(
        f'<p style="margin:4px 0">{line}</p>' for line in quotes_text.split("\n") if line.strip()
    )
    return f"""
<p>Hi {recipient_name or "there"},</p>

<p>Here's the travel insurance quote you requested from Ava at InsureHub.</p>

<p><b>Trip details:</b><br>{trip_summary}</p>

<p><b>Quote options:</b></p>
{quotes_html}

<p>Reply to this email if you have any questions — happy to help.</p>

<p>Best regards,<br>
<b>Ava, InsureHub Travel Insurance</b></p>
"""


def send_quotation_email(to_email: str, subject: str, body_html: str) -> bool:
    """Sends directly to the customer's own address (unlike the escalation
    path's Bcc-to-internal-agents trick) and carries no attachment. Returns
    False on any failure rather than raising — the caller decides how to
    surface that to the user."""
    sender, password = email_settings.get_sender(), email_settings.get_app_password()
    if not sender or not password:
        logger.warning(
            "Quotation email skipped — no sender/app password configured "
            "(Super Admin > Settings > Escalation Email Sender, or GMAIL_SENDER/GMAIL_APP_PASSWORD "
            "or OUTLOOK_SENDER/OUTLOOK_APP_PASSWORD in .env)."
        )
        return False

    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = sender
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(_email_wrapper(body_html, title="✈️ InsureHub — Your Travel Insurance Quote"), "html"))

        _smtp_send(sender, password, [to_email], msg)

        logger.info("Quotation email sent to %s", to_email)
        return True
    except Exception:
        logger.exception("Failed to send quotation email to %s", to_email)
        return False


def _email_wrapper(body: str, title: str = "🛡 InsureHub — Support Escalation") -> str:
    return f"""<!DOCTYPE html><html><body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
color:#1f2937;max-width:600px;margin:0 auto;padding:20px">
<div style="background:#4f46e5;color:white;padding:14px 20px;border-radius:8px 8px 0 0">
  <b>{title}</b>
</div>
<div style="background:#f9fafb;padding:20px;border:1px solid #e5e7eb;border-radius:0 0 8px 8px">
{body}
</div>
<p style="font-size:11px;color:#9ca3af;margin-top:16px">
This is an automated message from the InsureHub AI system.
</p>
</body></html>"""
